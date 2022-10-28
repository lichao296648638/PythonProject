import json
import pprint

import parsel
import random
import requests
from openpyxl import Workbook
import time
import re
from fake_useragent import UserAgent


# 创建一个 workbook
wb = Workbook()
# 获取被激活的 worksheet
ws = wb.active
# 设置单元格内容
ws['A1'] = '小区名字'
ws['B1'] = '所属分区'
ws['C1'] = '建成时间'
ws['D1'] = '房价m²'
ws['E1'] = '户数'
ws['F1'] = '物业公司'
ws['G1'] = '联系电话'

def random_sleep(mu, sigma):
    '''正态分布随机睡眠

    :param mu: 平均值

    :param sigma: 标准差，决定波动范围

    '''

    secs = random.normalvariate(mu, sigma)

    if secs <= 0:
        secs = mu  # 太小则重置为平均值

        time.sleep(secs)
# 爬取安居客
def ajk_data():

    # 安居客UA伪装
    ajk_headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
        'Cookie': 'ajkAuthTicket=TT=80c31316b82816d4925fa9c7ff60fa1c&TS=1666925129419&PBODY=H6zqLUbS9yWOBqhj3-E3YDm5DhtXce5Qh2Zf0chQJW5FqtWRDhLzkqkjMJLMV1bOeu_8KMFIEWX8GaodCGWI-5JijLE3RxSDz0YmPGOubg3Yir-okBUNRk-m2yzxb1m0eBL8smYpPwHPDchZ2eA6RrKwVLe2ixB-ditHtfpP3Dk&VER=2&CUID=GDzCNgmydFE6pPc8h58IvHKyO1Hvr8W-; Path=/; Domain=anjuke.com; Max-Age=86400; Expires=Sat, 29 Oct 2022 02:45:29 GMT; Secure; HttpOnly'
    }

    # 企业列表UA伪装
    company_list_headers = {
        'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/107.0.0.0 Safari/537.36',
        'cookie' : 'qcc_did=3bc29edc-d3e1-4a5b-9440-8739aabdfaf9; UM_distinctid=183162b782e17-06504d20da89a2-26021a51-1fa400-183162b782f18b4; acw_tc=77939ca316669237455222560efaf442b7ba1bb84ce50e9928e223a32b; CNZZDATA1254842228=349021138-1662523178-https%3A%2F%2Fwww.baidu.com%2F|1666923716; QCCSESSID=837663610f9b46ca5b4f315f32'
    }


    for page in range(7, 51):
            try:
                # 烟台安居客二手房列表
                ajk_list_url = f'https://yt.anjuke.com/community/p{page}/'
                # 发送请求
                response = requests.get(ajk_list_url, headers=ajk_headers)
                # 获取数据
                ajk_list_html = response.text
                print('--------------------------列表页爬取完成--------------------------------------------')

                # 解析数据 .li-row 的列表项
                list_selector = parsel.Selector(text=ajk_list_html)
                divs = list_selector.css('.li-row')
                if not divs:
                    random_sleep(70, 5)
                # 把div中的详情内容进行提取
                try:
                    for div in divs:
                        # 安居客详情页链接
                        item_url = div.css('::attr(href)').get()

                        # 抓取详情页信息
                        ajk_item_html = requests.get(url=item_url, headers=ajk_headers).text

                        # 解析详情页数据
                        item_selector = parsel.Selector(text=ajk_item_html)

                        # 户数
                        house_num = item_selector.css('.value.value_4::text').get().strip()

                        # 物业公司
                        property = item_selector.css('.info-list .column-1:nth-child(17) .value::text').get().strip()

                        # 查企业列表页链接
                        company_list_url = 'https://www.qcc.com/web/search'

                        # 企业查询请求参数
                        company_params = {
                            'key': property
                        }

                        property_tel = ''

                        if property is not '单位自管' or '业主自管' or '自主物业':
                            # 请求企业列表数据
                            company_list_html = requests.get(url=company_list_url, headers=company_list_headers,
                                                             params=company_params).text

                            print('---------------------企业表页--------------------')

                            company_select = parsel.Selector(text=company_list_html)
                            # 物业电话
                            property_tel = company_select.css('.phone-status+span::text').get()

                        # ---------写入单行数据-----------
                        # 小区名字
                        ws.append([div.css('.nowrap-min.li-community-title::text').get(),
                                   # 所属分区
                                   div.css('.split-line+span::text').get(),
                                   # 建成时间
                                   div.css('.year::text').get(),
                                   # 房价
                                   div.css('.community-price>strong::text').get(),
                                   # 户数
                                   house_num,
                                   # 物业公司
                                   property,
                                   # 物业电话
                                   property_tel
                                   ])
                        wb.save("test1.xlsx")
                        random_sleep(4, 1)
                        print('分区')
                        print(div.css('.split-line+span::text').get())
                        print('建成时间')
                        print(div.css('.year::text').get())
                        print('房价')
                        print(div.css('.community-price>strong::text').get())
                        print('户数')
                        print(house_num)
                        print('物业公司')
                        print(property)
                        print('物业电话')
                        print(property_tel)
                        print('------------------------------数据写入完成----------------------------------------')
                except:
                    continue
            except:
                continue



ajk_data()



