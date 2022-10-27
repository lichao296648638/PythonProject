from geopy.distance import geodesic
from openpyxl import Workbook
import openpyxl




# 获取网吧表格数据
def read_excel_bar():
    # 读表
    wb_ball = openpyxl.load_workbook('D:/app/GaodeEasyPoiPro2022/poi结果数据/Excel-烟台台球.xlsx')
    sheet_ball = wb_ball.get_sheet_by_name('poi数据')

    wb_bar = openpyxl.load_workbook('D:/app/GaodeEasyPoiPro2022/poi结果数据/Excel-烟台网吧.xlsx')
    sheet_bar = wb_bar.get_sheet_by_name('poi数据')

    #激活操作句柄
    ws = wb_bar.active

    # 设置网吧经纬度单元格数据来源
    cells_bar = sheet_bar['H2':'H348']
    # 遍历元组得到台球单元格内的经纬度数据数据
    bar_index = 1;
    ball_index = 1;
    for r_bar in cells_bar:
        # 检索下一个网吧,范围内台球厅计数清零
        dis_100 = 0
        dis_100_300 = 0
        dis_300_1000 = 0
        for c_bar in r_bar:
            #递增网吧检索下标
            bar_index += 1
            # 设置成计算距离适合的经纬度
            data_bar = c_bar.value.split(',')
            # 网吧经度
            longitude_bar = data_bar[1]
            # 网吧纬度
            latitude_bar = data_bar[0]

            # 设置台球经纬度单元格数据来源
            cells_ball = sheet_ball['H2':'H195']
            # 遍历元组得到网吧单元格内的经纬度数据数据
            for r_ball in cells_ball:
                for c_ball in r_ball:
                    # 设置成计算距离适合的经纬度，
                    data_ball = c_ball.value.split(',')
                    # 台球经度
                    longitude_ball = data_ball[1]
                    # 台球纬度
                    latitude_ball = data_ball[0]
                    # 计算两个坐标直线距离
                    dis = geodesic((longitude_bar, latitude_bar), (longitude_ball, latitude_ball)).m
                    if dis <= 100:
                        dis_100 += 1
                        ws[f'BF{bar_index + 1}'] = dis_100
                        wb_bar.save('test.xlsx')
                        print('100米范围台球厅' + str(dis_100) + '家')

                    if dis > 100 and dis <= 300:
                        dis_100_300 += 1
                        ws[f'BG{bar_index + 1}'] = dis_100_300
                        wb_bar.save('test.xlsx')
                        print('100-300米范围台球厅' + str(dis_100_300) + '家')

                    if dis > 300 and dis <= 1000:
                        dis_300_1000 += 1
                        ws[f'BH{bar_index + 1}'] = dis_300_1000
                        wb_bar.save('test.xlsx')
                        print('300-1000米范围台球厅' + str(dis_300_1000) + '家')


read_excel_bar()
#
#  # 输入的格式：纬度，经度
# print(geodesic((30.28708,120.12802999999997), (28.7427,115.86572000000001)).m)
# print(geodesic((30.28708,120.12802999999997), (28.7427,115.86572000000001)).km) #计算两个坐标直线距离